from flask import Flask, render_template, request, redirect, url_for, session, Blueprint
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re
import jwt



import logging
from logging.handlers import RotatingFileHandler
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

bp = Blueprint('main', __name__)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

log_file = 'app.log'
file_handler = RotatingFileHandler(log_file, maxBytes=1024*1024, backupCount=5, encoding="utf8")
file_handler.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

logger.addHandler(file_handler)

app = Flask(__name__)
app.secret_key = 'mysecretkey'

my_session = requests.Session()

system_url = "https://www.s-vfu.ru/user/rasp/new/"
server_url = "https://www.s-vfu.ru/user/rasp/new/ajax.php"

@bp.route('/')
@bp.route('/auth')
def index():

    return render_template('index.html')


@bp.route('/main', methods=['post', 'get'])
def authorize():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

    url = 'https://www.s-vfu.ru/?login=yes'

    data = {
        'AUTH_FORM': 'Y',
        'TYPE': 'AUTH',
        'USER_LOGIN': username,
        'USER_PASSWORD': password,
        'Login': ''
    }

    cookies = {
        "entersite": "www.s-vfu.ru",
    }

    res = my_session.post(url, data=data, cookies=cookies, verify=False)

    logger.info("Результат авторизации: " + str(res.status_code))

    right_index = res.text.find("<h1>") + 4
    left_index = res.text.find("</h1>")
    if left_index != -1:
        title = "Главная"
        name = res.text[right_index:left_index]
        logger.info(f'В систему вошел пользователь: {name}')
    else:
        title = "Login failed"
        right_index = res.text.find("<strong>Ошибка!</strong>")
        left_index = res.text.find("авторизация,") + 11
        name = "Ошибка авторизации! Неверные логин и/или пароль"
        logger.info(name)
        return redirect(url_for('index'))
        # return render_template('main.html')

    res = my_session.get(system_url)
    soup = BeautifulSoup(res.text, 'html.parser')
    buid = str(soup.find('input', {'name': 'buid'}))
    if buid is None:
        error = f'Ошибка при попытке получения идентификатора BITRIX пользователя {name}!'
        logger.error(error)
        return render_template('main.html', error=error)

    session['buid'] = buid

    return render_template('main.html', name=name, title=title)


@bp.route("/schedule", methods=['GET', 'POST'])
def schedule_parse():
    path = 'flask/static/'
    
    try:
        file = request.files['file']
    except KeyError as e:
        error = "Ошибка при получении файла: " + e
        logger.error(error)
        return render_template('main.html', error=error)
    
    try:
        file.save(path + file.filename)
    except PermissionError as e:
        error = "Ошибка про попытке сохранить файл с расписанием: " + str(e)
        logger.error(error)
        return render_template('main.html', error=error)
    except FileNotFoundError as e:
        error = "Ошибка про попытке сохранить файл с расписанием!" + str(e)
        logger.error(error)
        return render_template('main.html', error=error)


    form = request.form.get('form')
    fac = request.form.get('fac')

    if fac and form:
        buid = session.get('buid', 'ошибка')
        
        try:
            wb = load_workbook(filename=path + file.filename)
        except FileNotFoundError:
            error = "Указанный файл не найден!"
            logger.error(error)
            return render_template('main.html', error=error)

        sheets_names = wb.sheetnames

        # цикл по листам excel-файла
        for sh in sheets_names:
            
            if sh.find("курс") == -1:
                logger.warning(f'Возможно, название листа "{sh}" выбранной книги не соответствует требованиям формата файла с расписанием!')
                continue

            wb.active = sheets_names.index(sh)
            ws = wb.active

            course = str(ws.cell(row=2, column=1).value).strip()[0]
            
            year_and_semestr = str(ws.cell(row=1, column=1).value)
            year, semestr = get_year_and_semestr(year_and_semestr)

            if course == "None" and year_and_semestr == "None":
                logger.warning("Не удалось извлечь значение курса из файла!")
                logger.warning("Не удалось извлечь значение года и семестра из файла!")
                continue
            
            # цикл по всем группам
            for i in range(3, ws.max_column, 4):

                # получение названия группы
               
                if str(ws.cell(row=3, column=i).value).strip().lower() != "наименование группы":
                    continue

                group_name = str(ws.cell(row=4, column=i).value).strip()
                if group_name == "None":
                    logger.warning(f'Группа {group_name} была пропущена, т.к. не удалось извлечь название группы из файла! ')
                    continue

                if semestr == "1":
                    startdate = str(ws.cell(row=4, column=i + 2).value).strip() + "." + year
                    enddate = str(ws.cell(row=4, column=i + 3).value).strip() + "." + year
                else:
                    startdate = str(ws.cell(row=4, column=i + 2).value).strip() + "." + str(int(year) + 1)
                    enddate = str(ws.cell(row=4, column=i + 3).value).strip() + "." + str(int(year) + 1)

                # получения кода для формы обучения
                code = get_code(group_name)

                # получение списка подходящих групп
                response = query(id=buid, action="loadgroup", fac=fac, code=code,
                                    course=course, form=form, semestr=semestr, year=year)

                # получение нужной группы из списка
                try:
                    group = parse_loadgroup(response.text, group_name)
                except AttributeError as e:
                    logger.error(f"Ошибка при попытке получить нужную группу из списка: {e}.\nГруппа пропущена")
                    continue

                if group is None:
                    logger.error(f"Ошибка при попытке получить группу {group_name} из списка, полученного с сервера! Группа пропущена")
                    continue

                # получение id группы
                try:
                    group_id = group[group.find("|") + 1:]
                except AttributeError:
                    logger.error("Ошибка при попытке получить идентификатор группы! Группа была пропущена")
                    continue

                # получение РУПа
                response = query(id=buid, action="choicerup",
                                    fac=fac, course=course, form=form, semestr=semestr, year=year, groupname=group)

                full_semestr, filename = parse_choicerup(response.text)

                if full_semestr is None or filename is None:
                    logger.error("Ошибка при попытке получить значение семестра и РУПа группы! Группа была пропущена")
                    continue

                response = query(action="show", semestr=semestr, course=course, fac=fac,
                                    year=year, form=form, code=code, id_group=group_id, filename=filename)

                last_index = str(group_id).rfind("|") + 1

                full = f"{fac}|{filename}|{group_id[:last_index]}{full_semestr}|{course}|{year}|{semestr}|{group_id[last_index:len(group_id)]}|0{code}|{group_id[last_index:len(group_id)]}|{form[0]}"

                # получение ИД всех существующих занятий
                try:
                    lessons = get_lessons(response.text)
                except AttributeError as e:
                    logger.error(f"Ошибка при попытке получить идентификаторы существующих занятий: {e}.\nГруппа пропущена")
                    continue

                # # удаление всех существующих занятий
                # if len(lessons) > 0:
                #     for k in range(0, len(lessons), 4):
                #         query(action="delete", id=1,
                #               cell_id=lessons[k], full=full, fac=fac)
                #         query(action="remove", cell_id=lessons[k], full=full, id_group=group_id,
                #               filename=filename, semestr=semestr, full_semestr=full_semestr, course=course,
                #               fac=fac, year=year, form=form[0], code=code)
                
                weekday = str(
                                ws.cell(row=6, column=1).value).strip().upper()

                # цикл по занятиям одной группы
                for j in range(6, 42):

                    # проверка, что дисциплина есть (наличие пары)
                    if ws.cell(row=j, column=i).value is not None:
                        # получение дня недели
                        if j % 6 == 0:
                            weekday = str(
                                ws.cell(row=j, column=1).value).strip().upper()
            
                        time = str(ws.cell(row=j, column=2).value).replace(
                            ".", ":").replace(" -- ", "-")
                        
                        lesson_name = str(ws.cell(
                            row=j, column=i).value).split('\n')
                        
                        for q in range(len(lesson_name)):
                            lesson = lesson_name[q].strip()

                            podgruppa, lesson = get_podgruppa(lesson, q)
                            
                            try:
                                lesson, chet = get_parity(lesson)
                            except AttributeError as e:     
                                logger.error(f"Ошибка при попытке получить четность: {e}.\nГруппа пропущена")
                                continue

                            lecturers = str(
                                ws.cell(row=j, column=i + 1).value).split("\n")
                            
                            if len(lecturers) > 1:
                                lecturer_name = lecturers[q].strip()
                            else:
                                lecturer_name = lecturers[0].strip()

                            # добавление строки в таблицу
                            response = query(action="addrow", full=full)

                            # получение данных проподователя с сервера
                            if lecturer_name != "None":
                                hours, lecturer = parse_addrow(
                                    response.text, lecturer_name)
                            else: 
                                hours, lecturer = ("","")

                            activity = str(ws.cell(row=j, column=i + 2).value)
                            activity = get_activity(activity)

                            classrooms = str(ws.cell(row=j, column=i + 3).value).split("\n")

                            if len(classrooms) > 1:
                                classroom = classrooms[q].strip()
                            else:
                                classroom = classrooms[0].strip()

                            corpus, classroom = extract_corpus(classroom)

                            result = query(full=full, id_group=group_id, filename=filename, semestr=semestr, course=course,
                                        fac=fac, year=year, code=code, form=form[0], action="insertrow", full_semestr=full_semestr, lesson=lesson,
                                        lecturer=lecturer, weekday=weekday, time=time, chet=chet, startdate=startdate, enddate=enddate,
                                        activity=activity, corpus=corpus, classroom=classroom, hours=hours, podgruppa=podgruppa)

                    # response = query(action="public_check", full=full, fac=fac)
                    # if response.text.find(f'После нажатия кнопки "Применить" расписание группы ИМИ-{group_name} будет опубликовано') != -1:
                    #     print(response.text)

                    #     response = query(action="public", full=full, id_group=group_id, filename=filename, semestr=semestr,
                    #                 full_semestr=full_semestr, course=course, fac=fac, year=year, code=code, form=form[0])

        return render_template('schedule.html', buid=buid, res=result.text)
        # return redirect(system_url)

    else:
        error = "Ошибка при попытке получения факультета и формы обучения!"
        logger.error(error)
        return redirect(url_for('main'))


def get_lessons(html_code):
    # Найти таблицу по ее идентификатору
    soup = BeautifulSoup(html_code, 'html.parser')
    table = soup.find('table', id='mytable')

    # Получить все строки таблицы
    rows = table.find_all('tr')

    lessons_id = []

    # Пройтись по каждой строке и извлечь данные из ячеек
    for row in rows:
        cells = row.find_all('td')
        for cell in cells:
            cell_id = cell.get('id')
            if cell_id:
                lessons_id.append(cell_id)

    return lessons_id


def get_year_and_semestr(string):
    string1 = re.findall(r"\b\d+\b(?=\s*полугодие)", string)
    if string1:
        semestr = string1[0]
    else:
        semestr = None

    string2 = re.findall(r"полугодие\s+(\d+)\s*-\s*", string)
    if string2:
        year = string2[0][-4:]
    else:
        year = None
    return (year, semestr)


def get_podgruppa(lesson, q):
    if lesson.find("1/2") == -1:
        podgruppa = 0
    else:
        podgruppa = q + 1
    return (podgruppa, (lesson[:len(lesson) - 5]).strip())


def get_parity(lesson):
    if lesson.endswith("**"):
        return ((lesson[:len(lesson) - 2]).strip(), "2")
    elif lesson.endswith("*"):
        return ((lesson[:len(lesson) - 1]).strip(), "1")
    elif lesson.endswith("*,**") or lesson.endswith("*, **"):
        return ((lesson[:lesson.find("*")]).strip(), "0")
    else:
        return (lesson, "0")


def get_code(group_name):
    string = group_name.strip()[0].upper()
    if string == "Б":
        return "3"
    elif string == "М":
        return "4"
    elif string == "А":
        return "6"
    return "5"


def get_activity(act):
    act = act.replace(" ", "").replace('\\', "").replace('/', "").replace(',', '').replace("\n",'')
    act = act.lower()
    if act == "лек":
        return "лекция"
    elif act == "пр":
        return "практика"
    elif act == "лекпр" or act == 'леклаб':
        return "лекция, практика"
    elif act == "лаб":
        return "Лабораторная работа"
    elif act == "срс":
        return "самостоятельная работа"
    else:
        return ""


def extract_corpus(string):
     # Паттерн для поиска числа и слова
    pattern1 = r'[а-яА-ЯёЁ]+'
    pattern2 = r'\d+'
    # Ищем совпадения в строке
    word = (re.search(pattern1, string))
    number = (re.search(pattern2, string))
    if word is not None and number is not None:
        # Возвращаем слово из совпадения
        return (word.group(), number.group())
    elif string == "Спортивный":
        return ("Юность", string)
    else:
        return ("КФЕН", number.group())


def query(full="", id="", action="", fac="",
          code="", course="", form="",
          semestr="", year="", filename="",
          id_group="", groupname="",
          chet="", weekday="", activity="",
          corpus="", classroom="", lesson="",
          lecturer="", time="", full_semestr="", 
          startdate="", enddate="",
          cell_id="", hours="", podgruppa=""):

    url = "https://www.s-vfu.ru/user/rasp/new/ajax.php"

    # выбрать группу
    if action == 'loadgroup':

        data = {
            'id': id,
            'action': action,
            "fac": fac,
            "code": code,
            "course": course,
            "form": form,
            "semestr": semestr,
            "year": year
        }

    # выбрать руп
    elif action == 'choicerup':
        data = {
            'id': id,
            'action': action,
            "fac": fac,
            "course": course,
            "form": form,
            "semestr": semestr,
            "year": year,
            "groupname": groupname,
        }

    # выбор группы и РУПа
    elif action == 'show':
        url = "https://www.s-vfu.ru/user/rasp/new/"
        # headers = {
        #     "Content-Type": "multipart/form-data;"
        # }
        data = {
            "global_semestr": semestr,
            "semestr": full_semestr,
            "course": course,
            "fac": fac,
            "year": year,
            "formshort": form[0],
            "formname": form[2:],
            "action": action,
            "allplany": "on",
            "code": "0" + code,
            "id_group": id_group,
            "plan": filename,
            "startdate": startdate,
            "enddate": enddate
        }

    # удаление строки
    elif action == 'delete':

        data = {
            "action": action,
            "id": id,
            "data": cell_id,
            "full": full,
            "fac": fac
        }

    # подтверждение удаления строки
    elif action == 'remove':

        url = "https://www.s-vfu.ru/user/rasp/new/"

        data = {
            "data": full,
            "id_group": id_group,
            "filename": filename,
            "global_semestr": semestr,
            "semestr": full_semestr,
            "course": course,
            "fac": fac,
            "year": year,
            "form": "0" + code,
            "formshort": form,
            "action": "delete",
            "id": cell_id[2:]
        }

    # добавление строки
    elif action == 'addrow':
        id = 1
        data = {
            "action": action,
            "id": id,
            "full": full
        }

    elif action == "choicecorpus":
        data = {
            "id": id,
            "action": action,
            "corpus": corpus,
            "fac": fac
        }

    # вставка строки
    elif action == 'insertrow':

        url = "https://www.s-vfu.ru/user/rasp/new/"

        data = {
            "data": full,
            'id_group': id_group,
            "filename": filename,
            "global_semestr": semestr,
            "semestr": full_semestr,
            "course": course,
            "fac": fac,
            "year": year,
            "form": "0" + code,
            "formshort": form,
            'id': 1,
            'action': action,
            'I': lesson,
            "J": lecturer,
            "hours": hours,
            'podgruppa': podgruppa,
            "B": weekday,
            "F": time,
            "chet": chet,
            # "c": "09.01.2023",
            # "d": "30.06.2023",
            'c': startdate,
            'd': enddate,
            "H": activity,
            "L": corpus,
            "K": classroom
        }

    # публикация расписания
    elif action == 'public_check':

        data = {
            'id': 1,
            'action': "public",
            'full': full,
            "fac": fac
        }

    elif action == 'public':

        url = "https://www.s-vfu.ru/user/rasp/new/"

        data = {
            'data': full,
            'id_group': id_group,
            'filename': filename,
            'global_semestr': semestr,
            'semestr': full_semestr,
            'course': course,
            'fac': fac,
            'year': year,
            'form': "0" + code,
            'formshort': form,
            'action': action,
        }
    

    response = None

    try:
        response = my_session.post(url=url, data=data)
    except requests.exceptions.Timeout as e:
        logger.error("Произошла ошибка: " + e)
    except requests.exceptions.TooManyRedirects as e:
        logger.error("Произошла ошибка: " + e)
    except requests.exceptions.ConnectionError as e:
        logger.error("Произошла ошибка: " + e)
    except requests.exceptions.RequestException as e:
        logger.error("Произошла ошибка: ")
    except Exception as e:
        logger.error("Произошла ошибка: "+ e)

    if response is None or response.text == 'error':
        logger.warning(f'При попытке выполнить действие "{action}" произошла непредвиденная ошибка!')
        logger.info(f'Параметры запроса для действия "{action}": {data}')

    print(action)
    print(data)
    print("=============================================================\n\n")

    return response


def parse_loadgroup(html, groupname):
    soup = BeautifulSoup(html, 'html.parser')
    select = soup.find('select')
    if select:
        options = select.find_all('option')
        for option in options:
            value = option.get('value')
            if value and groupname in value:
                return value
    return None


def parse_choicerup(html):
    soup = BeautifulSoup(html, 'html.parser')
    plan = soup.find('input', {'name': 'plan'}).get('value')
    semestr = soup.find('input', {'name': 'semestr'}).get('value')
    if plan and semestr:
        return (semestr, plan) 
    return (None, None)


def parse_addrow(html, lecturer):
    soup = BeautifulSoup(html, 'html.parser')
    options = soup.find_all('option')

    surname, initials = lecturer.split()
    initials = initials.replace(".", "")

    for option in options:
        text = option.text
        if text.startswith(surname):
            string = text.split()
            lecturer_initials = string[1][0] + string[2][0]
            if initials == lecturer_initials:
                return ("", text + "|" + option['value'])

    else:
        # есть проблема совпадений по фамилии и инициалам а также полных тесок
        response = requests.get(
            url=f"https://www.s-vfu.ru/stud/searchadddata.php?tablename=svfudbnew.forexcel&term={surname} {initials[0]}")
        data = response.json()
        if len(data) > 0:
            logger.warning("Возможно, был указан неверный преподаватель, т.к. было найдено несколько совпадений по полученным фамилии и инициалам!")
        for d in data:
            string = str(d).split()
            if string[2].startswith(initials[1]):
                return (d[d.find("|") + 1:], "")

app.register_blueprint(bp, url_prefix='')

# HOST_PORT = "5000"
if __name__ == '__main__':
    app.debug = True
    app.run()
    # app.run(port=HOST_PORT)
