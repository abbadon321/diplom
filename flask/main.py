from openpyxl import load_workbook
import pandas as pd

wb = load_workbook(filename="flask\static\schedule.xlsx")
# sheets = []
sheets_names = wb.sheetnames
# for sh in sheets_names:
#     wb.active = sheets_names.index(sh)
#     sheet = wb.active
#     if sheet['B6'].value is not None:
#         print(sheet['B6'].value)
