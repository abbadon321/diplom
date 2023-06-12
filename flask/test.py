from bs4 import BeautifulSoup
import requests

from openpyxl import load_workbook
import re

# response = requests.get(
#             url=f"https://www.s-vfu.ru/stud/searchadddata.php?tablename=svfudbnew.forexcel&term=Попов В")
# data = response.json()

# print(data)

def get_podgruppa(lesson, q):
    if lesson.find("1/2") == -1:
        podgruppa = 0
    else:
        podgruppa = q + 1
    return (podgruppa, (lesson[:len(lesson) - 5]).strip())

wb = load_workbook(filename="C:/Users/Серега/Downloads/пример.xlsx")
sheets_names = wb.sheetnames
ws = wb.active

lesson_name = str(ws['C11'].value).split('\n')
# print(lesson_name)
for q in range(len(lesson_name)):
    lesson = lesson_name[q].strip()
    print(get_podgruppa(lesson, 1))




# print(str(ws['C7'].value))

# lesson = str(ws['C7'].value).split('\n')

# print(len(lesson))


# print(sheets_names)

# ws = wb.active

# for i in range(3, ws.max_column, 4):
#     group_name = str(ws.cell(row=4, column=i).value).strip()
#     if str(ws.cell(row=3, column=i).value).strip().lower() == "наименование группы":
#         for j in range(6, 42):
#             if ws.cell(row=j, column=i).value is not None:
#                 time = str(ws.cell(row=j, column=2).value).replace(
#                                     ".", ":").replace(" -- ", "-")
#                 lesson_name = str(ws.cell(
#                                     row=j, column=i).value).strip()
#                 print(time, lesson_name, end='\n\n')





# html = """
# <input type="hidden" name="data" value="ИМИ|010301_22_1ФИЦЭ.plx|7915|ИМИ-Б-М-22|2|1|2022|2|6581|03|6581|1"><input type="hidden" name="id_group" value="7915|ИМИ-Б-М-22|6581"><input type="hidden" name="filename" value="010301_22_1ФИЦЭ.plx"><input type="hidden" name="global_semestr" value="2"><input type="hidden" name="semestr" value="2"><input type="hidden" name="course" value="1"><input type="hidden" name="fac" value="ИМИ"><input type="hidden" name="year" value="2022"><input type="hidden" name="form" value="03"><input type="hidden" name="formshort" value="1"><div class="alert alert-danger">Внимание, у вас есть строки без преподавателей, если преподаватель не определен, очевидны ошибки при поиске свободных аудиторий и преподавателей</div><li>Иностранный язык (практика)-ПОНЕДЕЛЬНИК с 25.01 по 14.06 [08:00-09:35]</li><li>Иностранный язык (практика)-ВТОРНИК с 25.01 по 14.06 [08:00-09:35]</li><li>Иностранный язык (практика)-СРЕДА с 25.01 по 14.06 [08:00-09:35]</li><li>Иностранный язык (практика)-ЧЕТВЕРГ с 25.01 по 14.06 [08:00-09:35]</li><li>Иностранный язык (практика)-ПЯТНИЦА с 25.01 по 14.06 [08:00-09:35]</li><li>Алгебра (лекция)-ПЯТНИЦА с 09.01 по 30.06 [14:00-15:35]</li><li>Алгебра (практика)-ПЯТНИЦА с 09.01 по 30.06 [15:50-17:25]</li><li>Основы проектной деятельности (лекция, практика)-СУББОТА с 09.01 по 30.06 [09:50-11:25]</li><li>Основы проектной деятельности (лекция, практика)-СУББОТА с 09.01 по 30.06 [09:50-11:25]</li><li> Предупреждение поточная аудитория ВТОРНИК с 02.05 по 06.05 [11:40-13:15]-424(КФЕН) Местников С. В. читает Методы оптимизации в группе З-БП-ИВТ-18 у вас 424(КФЕН) - ВТОРНИК с 25.01 по 14.06 [11:40-13:15] читает Программирование </li><li> Предупреждение поточная аудитория СРЕДА с 09.01 по 30.06 [11:40-13:15]-Спортивный(Юность) Колодезникова М. Г. читает Элективные дисциплины по физической культуре и спорту в группе ГИ-С-ПР-22 у вас Спортивный(Юность) - СРЕДА с 09.01 по 30.06 [11:40-13:15] читает Элективные дисциплины по физической культуре и спорту </li><li> Предупреждение поточная аудитория СУББОТА с 09.01 по 30.06 [11:40-13:15]-Спортивный(Юность) Колодезникова М. Г. читает Элективные дисциплины по физической культуре и спорту в группе ГИ-С-ПР-22 у вас Спортивный(Юность) - СУББОТА с 09.01 по 30.06 [11:40-13:15] читает Элективные дисциплины по физической культуре и спорту </li><li> Предупреждение поточная аудитория ЧЕТВЕРГ с 27.03 по 31.03 [11:40-13:15]-353(КФЕН) Эверстова В. Н. читает Методика обучения математике в группе ИМИ-З-Б-ПОИМ-20 у вас 353(КФЕН) - ЧЕТВЕРГ с 09.01 по 30.06 [11:40-13:15] читает Аналитическая геометрия </li><li> Предупреждение поточная аудитория ЧЕТВЕРГ с 25.01 по 14.06 [11:40-13:15]-424(КФЕН) Акимов М. П. читает Введение в сквозные цифровые технологии в группе ИМИ-Б-ПМИ-22-1 у вас 424(КФЕН) - ЧЕТВЕРГ с 25.01 по 14.06 [11:40-13:15] читает Программирование </li>После нажатия кнопки "Применить" расписание группы ИМИ-Б-М-22 будет опубликовано<input type="hidden" name="action" value="public">
# """

# if html.find(f'После нажатия кнопки "Применить" расписание группы ИМИ-Б-М-22 будет опубликовано') != -1:
#     print("yes")


# html_code = """

# <!doctype html>
# <html>
#     <head>
#         <title>Северо-Восточный федеральный университет</title>
#         <meta http-equiv="X-UA-Compatible" content="IE=edge,chrome=1">
#         <meta charset="utf-8">
#         <meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=no">
#         <link href="/raspisanie/css/bootstrap.min.css" rel="stylesheet" media="screen">
#         <script src="/raspisanie/js/jquery.js"></script>
#         <script src="/raspisanie/js/bootstrap.min.js"></script>
#         <script src="//cdn.datatables.net/1.10.9/js/jquery.dataTables.min.js"  type="text/javascript"></script>
#         <link href="//cdn.datatables.net/1.10.9/css/jquery.dataTables.min.css" type="text/css"  data-template-style="true"  rel="stylesheet" />
#         <link rel="stylesheet" href="jquery-ui.min.css" type="text/css" />
#         <style>
#             .ui-autocomplete {z-index:2147483647; height: 150px; overflow-y: scroll; overflow-x: hidden;}
#         </style>
#         <script type="text/javascript" src="jquery-ui.min.js"></script>
#         <link href="context-menu-icons.eot" type="application/vnd.ms-fontobject" data-template-style="true"  rel="stylesheet" />
#         <link href="//cdnjs.cloudflare.com/ajax/libs/jquery-contextmenu/2.0.1/jquery.contextMenu.min.css" type="text/css"  data-template-style="true"  rel="stylesheet" />
#         <script src="//cdnjs.cloudflare.com/ajax/libs/jquery-contextmenu/2.0.1/jquery.contextMenu.min.js"></script>
#         <script src="//cdnjs.cloudflare.com/ajax/libs/jquery-contextmenu/2.0.1/jquery.ui.position.min.js"></script>
#         <script src="checkmobile.js"></script>
#         <style>
#             input.edit {
#                 width: 50%;
#                 height: 10%;
#                 font: bold 1.2em Arial, sans-serif;
#                 color: #0d3967;
#                 background-color: #cad5e2;
#                 text-shadow: #cad5ff 1px 1px 0;
#             }
#         </style>
#     </head>
#     <body>


# <h3>Редактор расписания (версия 2.0)</h3>
# <div class="navbar">
#     <div class="navbar-inner">

#         <a class="btn btn-navbar" data-toggle="collapse" data-target=".nav-collapse" href="#">
#             <span class="icon-bar"></span>
#             <span class="icon-bar"></span>
#             <span class="icon-bar"></span>
#         </a>
#         <a class="brand" href="/">s-vfu.ru</a>
#         <div class="nav-collapse">
#             <ul class="nav">
#                 <li><a href="#" id="Help">?</a></li>
#                  <li class="dropdown"><a class="dropdown-toggle"
#        data-toggle="dropdown" href="#">Расписание<b class="caret"></b>
#       </a>
#     <ul class="dropdown-menu">
#                 <li><a href="/user/rasp/video.php">Новости и видеоинструкции</a></li>
#                 <li><a href="#" id="Statistic">Статистика</a></li>
#                 <li><a href="#" id="Excel">Создать печатный вариант</a></li>
#                 <li><a href="/user/rasp/makeexcel.php?fac=ИМИ">Шахматка</a></li>
#                 <li><a href="/user/report/raspisanie/" target="_blank">Расписание всего университета</a></li>


#                 <li><a href="//s-vfu.ru/raspisanie/" target="_blank">s-vfu.ru/raspisanie</a></li>
#     </ul>
#   </li>
#       <li class="dropdown"><a class="dropdown-toggle"
#        data-toggle="dropdown" href="#">Почасовик<b class="caret"></b>
#       </a>
#     <ul class="dropdown-menu">
#                 <li><a href="#" id="Listhours">Список почасовиков</a></li>
# 				<li><a href="/user/ved/poisk.php" target="_blank">Добавить почасовика</a></li>
#                 <li><a href="#" id="Addhours">Добавить почасовика, если не найден в первых двух ссылках</a></li>
#                 </ul></li><li class="dropdown"><a class="dropdown-toggle"
#        data-toggle="dropdown" href="#">Доп.функционал<b class="caret"></b>
#       </a>
#     <ul class="dropdown-menu">
#                 <li><a href="#" id="AddCabinet">Справочник аудиторий</a></li>
#                 <li><a href="#" id="AddGroups">Работа с группами</a></li>
#                 <li><a href="/user/report/plany/" target="_blank">Рупы всего университета(для сверки)</a></li>
#                 <li><a href="/user/control/?o=group" target="_blank">Добавление группы для регистрации первокурсников ЭИОС</a></li>
#                 </ul></li>                <li><a href="#" id="Help"></a></li>
#                 <li><a href="#" id="Help"></a></li>
#                 <li><a href="#" id="Help"></a></li>
#                 <li><a href="#" id="Help"></a></li>
#             </ul>
#         </div><!-- /.nav-collapse -->
#     </div><!-- /navbar-inner -->
# </div><!-- /navbar -->
# <div class="alert alert-warning">
# <font color="red"><strong>Предварительные ГРУППЫ создаются в платформе <a href="/user/control/?o=group">Перейти в MyRoute</a></strong></font><br>
# <u>Добавлено копирование данных с группы на группу, <a href="/user/rasp/video.php">подробнее тут</a></u>. Внимание при смене года уже автоматически считает семестр - пригодится для дисциплин старых годов. Шаблон для работы с <a href="/user/rasp/macro.xla">Excel</a> <a href="/user/rasp/archive.php"> старым форматом ведомостей</a>
#     Весь функционал по работе с загрузке проверке Excel файлов перенесен в Доп.функции - Работа с Группами - фильтры влияют на вывод результата при выборе этого пункта</div>

# <input type="text" class="edit" name="url" value="/universitet/rukovodstvo-i-struktura/instituty/imi/Vremennoe_raspisanie_IMI_na_1_2020-2021/"><button id="updateurl">Обновить ссылку на первой странице сайта "Расписание"</button><br>УЧП:ИМИ            <form>
#                 <table>
#                     <tr><td>Учебный год:</td><td><select name="year">
# 								<option value="2022">2022/2023</option>
# 								<option value="2021">2021/2022</option>
#                                 <option value="2020">2020/2021</option>
#                                 <option value="2019">2019/2020</option>
#                                 <option value="2018">2018/2019</option>
# 								<option value="2017">2017/2018</option>
# 								<option value="2016">2016/2017</option>
# 								<option value="2023">2023/2024</option></select></td></tr>

#                     <tr><td><b>Курс:</b></td><td> <select name="course" id="course"><option value="1">1</option><option value="2">2</option><option value="3">3</option><option value="4">4</option><option value="5">5</option><option value="6">6</option><option value=99>1(предварительный)</option></select></td></tr>
#                     <tr><td>Форма обучения: </td><td><select name="form" id="studyform"><option value="1|очная">Очная</option><option value="7|очно-заочная">Очно-заочная</option><option value="2|заочная">заочная</option></select></td></tr>
#                     <tr><td>Уровень: </td><td><select name="code" id="code"><option value="3">Бакалавриат</option>
#                                 <option value="4">Магистратура</option>
#                                 <option value="5">Специалитет</option><option value="6">Аспирантура</option>                    </select></td></tr>
#                     <tr><td><b>Семестр*</b> </td><td><select name="semestr" id="semestr">
#                                 <option value="2">Весенний(09.01-30.06)</option>
# 							<option value="1">Зимний(01.09-31.12)</option>                            </select></td></tr>
#                 </table>
#                 <!--button type="button" class="btn btn-primary" data-toggle="modal" data-target="#myModal3" onclick="help()">?</button-->
#                 <!--button type="button" class="btn btn-primary" data-toggle="modal" data-target="#myModal3" onclick="listhours()">Справочник почасовиков своего факультета</button-->
#                 <!--button type="button" class="btn btn-primary" data-toggle="modal" data-target="#Modalrup" onclick="addhours('ИМИ')">Добавить почасовика</button-->
#                 <button type="button" class="btn btn-primary" onclick="loadgroup('loadgroup')">Шаг 1 Выбрать группы и план</button><br>

#             </form>
#             <input type="hidden" name="fac" value="ИМИ"><input type="hidden" name="buid" value="2902"><div id="show">---5----2022<h4>Заполнение расписания ИМИ для плана 090301_22П_5ИВТПО_zfo_vo.plx, учебной группы - З-БП-ИВТ-18, курс 5  семестр A ( учебный год 2022/2023)</h4><input type="hidden" name="data" value="ИМИ|090301_22П_5ИВТПО_zfo_vo.plx|6090|З-БП-ИВТ-18|A|5|2022|2|4703|03|4703|2"><br><button type="button" class="btn btn-primary"  data-toggle="modal" data-target="#Modalrup" onclick="Addrow()">Шаг 2. Добавить строку</button>
#       <button type="button" class="btn btn-primary"  data-toggle="modal" data-target="#Modalrup" onclick="Save()">Шаг 3. Сохранить (Публикация)</button><table id="mytable"  class="display" border=1>
#   <thead>
#   <tr><th>Когда</th><th>Кто и что</th><th>Где</th><th>Комментарии</th></tr></thead><tbody class="contentmenu"><tr><td id="A-696268">ПОНЕДЕЛЬНИК с 24.04 по 30.04 [08:00-09:35]</td><td  id="B-696268">Методы оптимизации (практика) -Местников С. В.</td><td id="C-696268">424(КФЕН)</td><td  id="D-696268">   </td></tr>
# <tr><td id="A-696249">ПОНЕДЕЛЬНИК с 17.04 по 23.04 [09:50-11:25]</td><td  id="B-696249">Теория принятия решений (лекция) -Местников С. В.</td><td id="C-696249">455(КФЕН)</td><td  id="D-696249">   </td></tr>
# <tr><td id="A-696269">ПОНЕДЕЛЬНИК с 24.04 по 30.04 [09:50-11:25]</td><td  id="B-696269">Методы оптимизации (практика) -Местников С. В.</td><td id="C-696269">455(КФЕН)</td><td  id="D-696269">   </td></tr>
# <tr><td id="A-696259">ПОНЕДЕЛЬНИК с 17.04 по 30.04 [14:00-15:35]</td><td  id="B-696259">Защита интеллектуальной собственности (лекция, практика) -Ефремова Е. А.</td><td id="C-696259">328(КФЕН)</td><td  id="D-696259">   </td></tr>
# <tr><td id="A-696260">ПОНЕДЕЛЬНИК с 17.04 по 30.04 [15:50-17:25]</td><td  id="B-696260">Защита интеллектуальной собственности (лекция, практика) -Ефремова Е. А.</td><td id="C-696260">540(КФЕН)</td><td  id="D-696260">   </td></tr>
# <tr><td id="A-696261">ПОНЕДЕЛЬНИК с 17.04 по 30.04 [17:40-19:15]</td><td  id="B-696261">Защита интеллектуальной собственности (лекция, практика) -Ефремова Е. А.</td><td id="C-696261">540(КФЕН)</td><td  id="D-696261">   </td></tr>
# <tr><td id="A-696253">ВТОРНИК с 17.04 по 23.04 [09:50-11:25]</td><td  id="B-696253">Теория принятия решений (практика) -Местников С. В.</td><td id="C-696253">540(КФЕН)</td><td  id="D-696253">   </td></tr>
# <tr><td id="A-696244">ВТОРНИК с 24.04 по 30.04 [09:50-11:25]</td><td  id="B-696244">Организация и технологии защиты информации (практика) -Тимофеев Е. М.</td><td id="C-696244">540(КФЕН)</td><td  id="D-696244">   </td></tr>
# <tr><td id="A-696270">ВТОРНИК с 02.05 по 06.05 [09:50-11:25]</td><td  id="B-696270">Методы оптимизации (практика) -Местников С. В.</td><td id="C-696270">424(КФЕН)</td><td  id="D-696270">   </td></tr>
# <tr><td id="A-696245">ВТОРНИК с 24.04 по 30.04 [11:40-13:15]</td><td  id="B-696245">Организация и технологии защиты информации (практика) -Тимофеев Е. М.</td><td id="C-696245">540(КФЕН)</td><td  id="D-696245">   </td></tr>
# <tr><td id="A-696252">ВТОРНИК с 17.04 по 23.04 [11:40-13:15]</td><td  id="B-696252">Теория принятия решений (практика) -Местников С. В.</td><td id="C-696252">540(КФЕН)</td><td  id="D-696252">   </td></tr>
# <tr><td id="A-696271">ВТОРНИК с 02.05 по 06.05 [11:40-13:15]</td><td  id="B-696271">Методы оптимизации (зачет) -Местников С. В.</td><td id="C-696271">424(КФЕН)</td><td  id="D-696271">   </td></tr>
# <tr><td id="A-696262">ВТОРНИК с 17.04 по 23.04 [14:00-15:35]</td><td  id="B-696262">Защита интеллектуальной собственности (практика) -Ефремова Е. А.</td><td id="C-696262">324(КФЕН)</td><td  id="D-696262">   </td></tr>
# <tr><td id="A-696275">ВТОРНИК с 24.04 по 08.05 [14:00-15:35]</td><td  id="B-696275">Организация и планирование производства (лекция, практика) -Ефремова Е. А.</td><td id="C-696275">324(КФЕН)</td><td  id="D-696275">   </td></tr>
# <tr><td id="A-696263">ВТОРНИК с 17.04 по 23.04 [15:50-17:25]</td><td  id="B-696263">Защита интеллектуальной собственности (практика) -Ефремова Е. А.</td><td id="C-696263">324(КФЕН)</td><td  id="D-696263">   </td></tr>
# <tr><td id="A-696276">ВТОРНИК с 24.04 по 08.05 [15:50-17:25]</td><td  id="B-696276">Организация и планирование производства (лекция, практика) -Ефремова Е. А.</td><td id="C-696276">455(КФЕН)</td><td  id="D-696276">   </td></tr>
# <tr><td id="A-696250">СРЕДА с 17.04 по 23.04 [09:50-11:25]</td><td  id="B-696250">Теория принятия решений (лекция) -Местников С. В.</td><td id="C-696250">457(КФЕН)</td><td  id="D-696250">   </td></tr>
# <tr><td id="A-696254">СРЕДА с 17.04 по 23.04 [11:40-13:15]</td><td  id="B-696254">Теория принятия решений (практика) -Местников С. В.</td><td id="C-696254">426(КФЕН)</td><td  id="D-696254">   </td></tr>
# <tr><td id="A-696264">СРЕДА с 17.04 по 23.04 [14:00-15:35]</td><td  id="B-696264">Методы оптимизации (лекция) -Местников С. В.</td><td id="C-696264">457(КФЕН)</td><td  id="D-696264">   </td></tr>
# <tr><td id="A-696277">СРЕДА с 24.04 по 08.05 [14:00-15:35]</td><td  id="B-696277">Организация и планирование производства (лекция, практика) -Ефремова Е. А.</td><td id="C-696277">540(КФЕН)</td><td  id="D-696277">   </td></tr>
# <tr><td id="A-696265">СРЕДА с 17.04 по 23.04 [15:50-17:25]</td><td  id="B-696265">Методы оптимизации (лекция) -Местников С. В.</td><td id="C-696265">457(КФЕН)</td><td  id="D-696265">   </td></tr>
# <tr><td id="A-696278">СРЕДА с 24.04 по 08.05 [15:50-17:25]</td><td  id="B-696278">Организация и планирование производства (лекция, практика) -Ефремова Е. А.</td><td id="C-696278">540(КФЕН)</td><td  id="D-696278">   </td></tr>
# <tr><td id="A-696246">ЧЕТВЕРГ с 24.04 по 30.04 [09:50-11:25]</td><td  id="B-696246">Организация и технологии защиты информации (зачет) -Тимофеев Е. М.</td><td id="C-696246">445(КФЕН)</td><td  id="D-696246">   </td></tr>
# <tr><td id="A-696255">ЧЕТВЕРГ с 17.04 по 23.04 [09:50-11:25]</td><td  id="B-696255">Теория принятия решений (практика) -Местников С. В.</td><td id="C-696255">457(КФЕН)</td><td  id="D-696255">   </td></tr>
# <tr><td id="A-696256">ЧЕТВЕРГ с 17.04 по 23.04 [11:40-13:15]</td><td  id="B-696256">Теория принятия решений (зачет) -Местников С. В.</td><td id="C-696256">551(КФЕН)</td><td  id="D-696256">   </td></tr>
# <tr><td id="A-696266">ЧЕТВЕРГ с 17.04 по 23.04 [14:00-15:35]</td><td  id="B-696266">Методы оптимизации (лекция) -Местников С. В.</td><td id="C-696266">445(КФЕН)</td><td  id="D-696266">   </td></tr>
# <tr><td id="A-696267">ЧЕТВЕРГ с 17.04 по 23.04 [15:50-17:25]</td><td  id="B-696267">Методы оптимизации (практика) -Местников С. В.</td><td id="C-696267">445(КФЕН)</td><td  id="D-696267">   </td></tr>
# <tr><td id="A-696282">ЧЕТВЕРГ с 24.04 по 08.05 [15:50-17:25]</td><td  id="B-696282">Основы предпринимательства (лекция, практика) -Ефремова Е. А.</td><td id="C-696282">536(КФЕН)</td><td  id="D-696282">   </td></tr>
# <tr><td id="A-696279">ЧЕТВЕРГ с 24.04 по 08.05 [17:40-19:15]</td><td  id="B-696279">Основы предпринимательства (лекция, практика) -Ефремова Е. А.</td><td id="C-696279">536(КФЕН)</td><td  id="D-696279">   </td></tr>
# <tr><td id="A-696233">ПЯТНИЦА с 14.04 по 30.04 [09:50-11:25]</td><td  id="B-696233">Нейросетевые технологии (лекция) -Калачикова У. С.</td><td id="C-696233">435(КФЕН)</td><td  id="D-696233">   </td></tr>
# <tr><td id="A-696234">ПЯТНИЦА с 14.04 по 23.04 [11:40-13:15]</td><td  id="B-696234">Нейросетевые технологии (Лабораторная работа) -Калачикова У. С.</td><td id="C-696234">435(КФЕН)</td><td  id="D-696234">   </td></tr>
# <tr><td id="A-696240">ПЯТНИЦА с 14.04 по 23.04 [11:40-13:15]</td><td  id="B-696240">Нейросетевые технологии (Лабораторная работа) -Калачикова У. С.</td><td id="C-696240">436(КФЕН)</td><td  id="D-696240">   </td></tr>
# <tr><td id="A-696238">ПЯТНИЦА с 14.04 по 23.04 [14:00-15:35]</td><td  id="B-696238">Организация и технологии защиты информации (лекция, практика) -Тимофеев Е. М.</td><td id="C-696238">536(КФЕН)</td><td  id="D-696238">   </td></tr>
# <tr><td id="A-696242">ПЯТНИЦА с 14.04 по 23.04 [15:50-17:25]</td><td  id="B-696242">Организация и технологии защиты информации (лекция, практика) -Тимофеев Е. М.</td><td id="C-696242">536(КФЕН)</td><td  id="D-696242">   </td></tr>
# <tr><td id="A-696281">ПЯТНИЦА с 24.04 по 08.05 [15:50-17:25]</td><td  id="B-696281">Основы предпринимательства (лекция, практика) -Ефремова Е. А.</td><td id="C-696281">536(КФЕН)</td><td  id="D-696281">   </td></tr>
# <tr><td id="A-696280">ПЯТНИЦА с 24.04 по 08.05 [17:40-19:15]</td><td  id="B-696280">Основы предпринимательства (лекция, практика) -Ефремова Е. А.</td><td id="C-696280">536(КФЕН)</td><td  id="D-696280">   </td></tr>
# <tr><td id="A-696272">СУББОТА с 17.04 по 08.05 [08:00-09:35]</td><td  id="B-696272">Интернет-предпринимательство (лекция, практика) -Ефремова Е. А.</td><td id="C-696272">519(КФЕН)</td><td  id="D-696272">   </td></tr>
# <tr><td id="A-696243">СУББОТА с 14.04 по 16.04 [09:50-11:25]</td><td  id="B-696243">Организация и технологии защиты информации (лекция, практика) -Тимофеев Е. М.</td><td id="C-696243">536(КФЕН)</td><td  id="D-696243">   </td></tr>
# <tr><td id="A-696273">СУББОТА с 17.04 по 08.05 [09:50-11:25]</td><td  id="B-696273">Интернет-предпринимательство (лекция, практика) -Ефремова Е. А.</td><td id="C-696273">519(КФЕН)</td><td  id="D-696273">   </td></tr>
# <tr><td id="A-696274">СУББОТА с 17.04 по 30.04 [11:40-13:15]</td><td  id="B-696274">Интернет-предпринимательство (лекция, практика) -Ефремова Е. А.</td><td id="C-696274">519(КФЕН)</td><td  id="D-696274">   </td></tr>
# <tr><td id="A-696235">СУББОТА с 14.04 по 23.04 [14:00-15:35]</td><td  id="B-696235">Нейросетевые технологии (лекция) -Калачикова У. С.</td><td id="C-696235">424(КФЕН)</td><td  id="D-696235">   </td></tr>
# <tr><td id="A-696236">СУББОТА с 14.04 по 16.04 [15:50-17:25]</td><td  id="B-696236">Нейросетевые технологии (Лабораторная работа) -Калачикова У. С.</td><td id="C-696236">424(КФЕН)</td><td  id="D-696236">   </td></tr>
# <tr><td id="A-696239">СУББОТА с 14.04 по 16.04 [15:50-17:25]</td><td  id="B-696239">Нейросетевые технологии (Лабораторная работа) -Калачикова У. С.</td><td id="C-696239">436(КФЕН)</td><td  id="D-696239">   </td></tr>
# <tr><td id="A-696237">СУББОТА с 14.04 по 16.04 [17:40-19:15]</td><td  id="B-696237">Нейросетевые технологии (Лабораторная работа) -Калачикова У. С.</td><td id="C-696237">424(КФЕН)</td><td  id="D-696237">   </td></tr>
# <tr><td id="A-696241">СУББОТА с 14.04 по 16.04 [17:40-19:15]</td><td  id="B-696241">Нейросетевые технологии (Лабораторная работа) -Калачикова У. С.</td><td id="C-696241">436(КФЕН)</td><td  id="D-696241">   </td></tr>
# </tbody>
#       </table>
#       <button type="button" class="btn btn-primary" data-toggle="modal" data-target="#Modalrup" onclick="Addrow()">Шаг 2. Добавить строку</button>
#       <button type="button" class="btn btn-primary" data-toggle="modal"  data-target="#Modalrup" onclick="Save()">Шаг 3.Сохранить (Публикация)</button>
#           </div>
#     <div class="modal fade" id="myModal3" tabindex="-1" role="dialog"  aria-hidden="true">
#         <div class="modal-dialog" role="document">
#             <div class="modal-content">
#                 <div class="modal-header">
#                     <button type="button" class="close" data-dismiss="modal" aria-label="Close">
#                         <span aria-hidden="true">&times;</span>
#                     </button>
#                 </div>
#                 <div class="modal-body" id="rup">Подождите.....
#                 </div>
#                 <div class="modal-header">
#                     <button type="button" class="btn btn-secondary" data-dismiss="modal">Закрыть</button>
#                 </div>
#             </div>
#         </div>
#     </div>
#     <div class="modal fade" id="Modalrup" tabindex="-1" role="dialog"  aria-hidden="true">
#         <div class="modal-dialog" role="document">
#             <div class="modal-content">

#                 <div class="modal-header">
#                     <button type="button" class="close" data-dismiss="modal" aria-label="Close">
#                         <span aria-hidden="true">&times;</span>
#                     </button>
#                 </div>
#                 <form method="POST" enctype="multipart/form-data">
#                     <div class="modal-body" id="rup2">Подождите.....
#                     </div>
#                     <div class="modal-header">
#                         <input type="submit" value="Применить">
#                         <button type="button" class="btn btn-secondary" data-dismiss="modal">Закрыть</button>
#                     </div>
#                 </form>
#             </div>
#         </div>
#     </div>
#     <script type="text/javascript" src="library.js?new"></script>
# </body>
# </html>


# """

# soup = BeautifulSoup(html_code, 'html.parser')

# # Найти таблицу по ее идентификатору
# table = soup.find('table', id='mytable')

# # Получить все строки таблицы
# rows = table.find_all('tr')
# data = []
# # Пройтись по каждой строке и извлечь данные из ячеек
# for row in rows:
#     cells = row.find_all('td')
#     for cell in cells:
#         cell_id = cell.get('id')
#         if cell_id:
#             data.append(cell_id)

# print(data)


# group = "02.03.02|7471|ИМИ-Б-ФИИТ-21|5998"
# group_id = group[group.find("|") + 1:]

# semestr = "2"
# course = "2"
# fac = "ИМИ"
# filename = "02030201_22_2ФИИТ.plx"
# code = "3"
# year = "2022"
# form = "1"

# last_index = str(group_id).rfind("|") + 1
# full_semestr = str((int(course) - 1) * 2 + int(semestr))
# full = full = f"{fac}|{filename}|{group_id[:last_index]}{full_semestr}|{course}|{year}|{semestr}|{group_id[last_index:len(group_id)]}|0{code}|{group_id[last_index:len(group_id)]}|{form[0]}"

# print(full)

# my_session = requests.Session()

# url = 'https://www.s-vfu.ru/?login=yes'


# data = {
#     'AUTH_FORM': 'Y',
#     'TYPE': 'AUTH',
#     'USER_LOGIN': "rom.na",
#     'USER_PASSWORD': "CfvjqkjdFY1937",
#     'Login': ''
# }

# cookies = {
#     "entersite": "www.s-vfu.ru",
# }

# res = my_session.post(url, data=data, cookies=cookies, verify=False)

# my_cookies = res.cookies

# # res = my_session.post(url="https://www.s-vfu.ru/user/rasp/new/ajax.php",
# #                       data={
# #                           "action": "aboutgroup",
# #                           "id": "2902"
# #                       }
# #                       )


# # def parse_loadgroup(html, groupname):
# #     soup = BeautifulSoup(html, 'html.parser')
# #     select = soup.find('select')
# #     if select:
# #         options = select.find_all('option')
# #         for option in options:
# #             value = option.get('value')
# #             if value and groupname in value:
# #                 return value
# #     return None


# def query(id=None, action=None, fac=None,
#           code=None, course=None, form=None,
#           semestr=None, year=None, filename=None,
#           id_group=None, groupname=None, full=None,
#           chet=None, weekday=None, activity=None,
#           corpus=None, classroom=None, lesson=None,
#           lecturer=None, time=None, cell_id=None):

#     url = "https://www.s-vfu.ru/user/rasp/new/ajax.php"

#     # добавление строки
#     if action == 'addrow':
#         id = 1
#         # groupname: 02.03.02|7471|ИМИ-Б-ФИИТ-21|5998
#         # ИМИ|02030201_22_2ФИИТ.plx|7471|ИМИ-Б-ФИИТ-21|3|2|2022|1|5998|03|5998|1
#         data = {"id": id,
#                 "full": full
#                 }

#     # удаление строки
#     elif action == 'delete':

#         data = {
#             "action": action,
#             "id": id,
#             "data": cell_id,
#             "full": full,
#             "fac": fac
#         }

#     # подтверждение удаления строки
#     elif action == 'remove':

#         url = "https://www.s-vfu.ru/user/rasp/new/"

#         data = {
#             "data": full,
#             "id_group": id_group,
#             "filename": filename,
#             "global_semestr": semestr,
#             "semestr": "A",
#             "course": course,
#             "fac": fac,
#             "year": year,
#             "form": "0" + code,
#             "formshort": form,
#             "action": "delete",
#             "id": cell_id[2:]
#         }

#     # вставка строки
#     elif action == 'insertrow':

#         data = {
#             "data": full,
#             'courseequalsemestr': 0,
#             'id_group': id_group,
#             "filename": filename,
#             "global_semestr": semestr,
#             "semestr": (course-1) * 2 + semestr,
#             "course": course,
#             "fac": fac,
#             "year": year,
#             "form": groupname[3:5],
#             "formshort": 1,
#             'id': 1,
#             'action': action,
#             'I': lesson,
#             # "Акинин Михаил Александрович|895035670"
#             "J": lecturer,
#             "hours": lecturer[lecturer.find("|") + 1:],
#             'poggruppa': 0,
#             "B": weekday,
#             "F": time,
#             "chet": chet,
#             "c": "09.01.2023",
#             "d": "30.06.2023",
#             "H": activity,
#             "L": corpus,
#             "K": classroom
#         }

#     # публикация расписания
#     elif action == 'public1':

#         data = {
#             'id': id,
#             'action': action,
#             'full': full,
#             "fac": fac
#         }
#     elif action == 'public2':

#         data = {
#             'data': fac + "|" + filename + "|" + id_group + "|" + groupname[:last_index] + "|" + str((course-1) * 2 + semestr) + "|" + course + "|" + year + "|" + semestr + "|" + groupname[last_index:len(groupname)] + "|" + groupname[3:5] + "|" + groupname[last_index:len(groupname)] + "|" + form,
#             'id_group': id_group,
#             'filename': filename,
#             'global_semestr': semestr,
#             'semestr': (course-1) * 2 + semestr,
#             'course': course,
#             'fac': fac,
#             'year': year,
#             'form': groupname[3:5],
#             'formshort': form[0],
#             'action': action,
#         }

#     # сохранение расписания
#     elif action == 'apply':

#         data = {
#             'id': id,
#             'action': action,
#             'filename': filename,
#             "course": course,
#             "id_group": id_group,
#             "semestr": semestr,
#             "year": year,
#             "fac": fac
#         }

#     # выбрать группу
#     elif action == 'loadgroup':

#         data = {
#             'id': id,
#             'action': action,
#             "fac": fac,
#             "code": code,
#             "course": course,
#             "form": form,
#             "semestr": semestr,
#             "year": year
#         }

#     # выбрать руп
#     elif action == 'choicerup':

#         data = {
#             'id': id,
#             'action': action,
#             "fac": fac,
#             "course": course,
#             "form": form,
#             "semestr": semestr,
#             "year": year,
#             "groupname": groupname,
#         }

#     response = my_session.post(
#         url=url, data=data, cookies=my_cookies)
#     return response


# response = query(action="delete", id=1,
#                  cell_id="A-701219", full="ИМИ|090301_22П_5ИВТПО_zfo_vo.plx|6090|З-БП-ИВТ-18|A|5|2022|2|4703||4703|2", fac="ИМИ")

# print(response.text)

# print("\n\n")

# response = query(action="remove", cell_id="A-701219", full="ИМИ|090301_22П_5ИВТПО_zfo_vo.plx|6090|З-БП-ИВТ-18|A|5|2022|2|4703||4703|2",
#                  id_group="6090|З-БП-ИВТ-18|4703",
#                  filename="090301_22П_5ИВТПО_zfo_vo.plx", semestr="2", course="5",
#                  fac="ИМИ", year="2022", form="2", code="3")

# print(response.text)

# response = query(2902, "loadgroup", "ИМИ", 3, 1,
#                  "1|очная", 2, 2022)

# print(type(parse_loadgroup(response, "Б-М-22")))

# lesson = {
#     "ИД группы": group_id,
#     "номер пары": j - 5,
#     "день недели": weekday,
#     "временной отрезок": time,
#     "название дисциплины": lesson_name,
#     "ФИО преподавателя": lecturer,
#     "вид деятельности": activity,
#     "номер аудитории": classroom,
# }
# schedule.setdefault(
#     group_name, []).append(lesson)
# print(group_id, filename, semestr, course,
#       fac, form[0], lesson_name,
#       lecturer, weekday, time, chet,
#       activity, corpus, classroom, year, sep="\n")

# query(action="choicecorpus",
#       id=99999, corpus=corpus, fac=0)

# print(response)

# group = "02.03.02|7471|ИМИ-Б-ФИИТ-21|5998"
# print(group[3:5])

# def parse_addrow(html, lecturer):
#     soup = BeautifulSoup(html, 'html.parser')
#     options = soup.find_all('option')

#     surname, initials = lecturer.split()
#     initials = initials.replace(".", "")

#     for option in options:
#         text = option.text
#         if text.startswith(surname):
#             string = text.split()
#             lecturer_initials = string[1][0] + string[2][0]
#             if initials == lecturer_initials:
#                 return text + "|" + option['value']

#     else:
#         # есть проблема совпадений по фамилии и инициалам а также полных тесок
#         response = requests.get(
#             url=f"https://www.s-vfu.ru/stud/searchadddata.php?tablename=svfudbnew.forexcel&term={surname} {initials[0]}")
#         data = response.json()
#         for d in data:
#             string = d.split()
#             if string[2].startswith(initials[1]):
#                 return d
#             else: "Преподаватель не найден!"


# html = '<select name="hours"><option value=""></option><option value="895038073">Акимов Федор Револьевич</option><option value="714069">Алексеев Николай Кириллович</option><option value="90258224">Божевольная Зоя Анатольевна</option><option value="895038074">Варламова Анастасия Гаврииловна</option><option value="895035721">Васильева Лира Петровна</option><option value="895035612">Габышева Анна Михайловна</option><option value="895038096">Герасимов Георгий Егорович</option><option value="895038199">Голоков Вячеслав Валерьевич</option><option value="717477">Дедюкина Любовь Лукинична</option><option value="717689">Донская Маргарита Ивановна</option><option value="718121">Егорова Валентина Никифоровна</option><option value="219536635">Ефимова Кристина Семеновна</option><option value="718509">Жафяров Акрям Жафярович</option></select>'
# lecturer = "Попов В.В."

# print(parse_addrow(html, lecturer))

# lecturer = "Акинин Михаил Александрович|895035670"
# print(lecturer[lecturer.find("|") + 1:])


# wb = load_workbook(filename='C:\\Users\\user\\Documents\\GitHub\\diplom\\flask\\static\\тест.xlsx')
# sheets_names = wb.sheetnames
# wb.active = sheets_names.index("1 курс_МО")
# ws = wb.active
# print(ws.title)
# text = str(ws.cell(row=27, column=4).value).split("\n")
# print(text)
# s = None
# if str(s) == "None":
#     print("yes")


# def extract_word(string):
#     # Паттерн для поиска числа и слова
#     pattern = r'\b(\d+)\b\s+([a-zA-Zа-яА-Я]{2,})\b'

#     # Ищем совпадения в строке
#     match = re.search(pattern, string)

#     if match is not None:
#         # Возвращаем слово из совпадения
#         return match.group(2)
#     else:
#         return None

# # цикл по листам excel-файла
# for sh in sheets_names:
#     wb.active = sheets_names.index(sh)
#     ws = wb.active
#     print("\n", ws.title)
#     text = str(ws.cell(row=3, column=6).value).split("\n")
#     print(text)
#     for i in range(6, ws.max_column, 4):
#         for j in range(6, 42):
#             if ws.cell(row=j, column=i).value is not None:
#                 string = str(ws.cell(row=j, column=i).value)
#                 result = extract_word(string) if (result := extract_word(string)) is not None else "КФЕН"
#                 print(result, end=", ")

# if str(ws.cell(row=j, column=i).value).find("*,**") != -1:
# if str(ws.cell(row=j, column=i).value).find("*") != -1:
#     print(ws.cell(row=j, column=i).value)
#     print( (ws.cell(row=j, column=i).value.strip()))
#     print()
#     # print((ws.cell(row=j, column=i).value.strip()).split())
# print(j, " ",(ws.cell(row=j, column=i).value.strip()))


# surname = "Попов"
# initials = "ВВ"
# response = requests.get(
#     url=f"https://www.s-vfu.ru/stud/searchadddata.php?tablename=svfudbnew.forexcel&term={surname} {initials[0]}")
# data = response.json()
# # print(data)
# for d in data:
#     string = d.split()
#     if string[2].startswith(initials[1]):
#         print(d)


# def find_value_with_substring(html, substring):
#     soup = BeautifulSoup(html, 'html.parser')
#     select = soup.find('select')
#     if select:
#         options = select.find_all('option')
#         for option in options:
#             value = option.get('value')
#             if value and substring in value:
#                 return value
#     return None


# substring = "Б-М-21"
# html = 'Семестр 1<hr><select name="groupname"><optgroup label="Есть расписание"><option value="09.03.01|7618|ИМИ-Б-ИВТ-21-1|5954">(18.10 15:32) - ИМИ-Б-ИВТ-21-1(09.03.01-Технологии разработки программного обеспечения) -4 г. (20) </option><option value="01.03.01|7468|ИМИ-Б-М-21|5996">(24.10 12:09) - ИМИ-Б-М-21(01.03.01-Математика) -4 г. (10) </option><option value="02.03.02|7471|ИМИ-Б-ФИИТ-21|5998">(19.09 11:33) - ИМИ-Б-ФИИТ-21(02.03.02-Фундаментальная информатика и информационные технологии) -4 г. (21) </option><option value="09.03.01|7619|ИМИ-Б-ИВТ-21-2|5999">(19.09 11:17) - ИМИ-Б-ИВТ-21-2(09.03.01-Технологии разработки программного обеспечения) -4 г. (21) </option><option value="11.03.02|7467|ИМИ-Б-ИТСС-21|6003">(16.09 11:30) - ИМИ-Б-ИТСС-21(11.03.02-Инфокоммуникационные технологии и системы связи) -4 г. (18) </option><option value="44.03.01|7469|ИМИ-Б-МПО-21|6004">(15.09 12:51) - ИМИ-Б-МПО-21(44.03.01-Математика) -4 г. (13) </option><option value="44.03.05|7470|ИМИ-Б-ПОИМ-21|6005">(18.10 15:16) - ИМИ-Б-ПОИМ-21(44.03.05-Информатика и математика) -5 л. (13) </option><option value="09.03.03|7856|ИМИ-Б-ПИГМУ-21|6406">(27.09 14:13) - ИМИ-Б-ПИГМУ-21(09.03.03-Прикладная информатика в государственном и муниципальном управлении) -4 г. (18) </option><option value="09.03.03|7855|ИМИ-Б-ПИЭ-21|6407">(27.09 14:14) - ИМИ-Б-ПИЭ-21(09.03.03-Прикладная информатика в экономике) -4 г. (19) </option><option value="01.03.02|7623|ИМИ-Б-ПМИ-21|6724">(24.10 12:10) - ИМИ-Б-ПМИ-21(01.03.02-Прикладная математика и информатика) -4 г. (28) </option></select><button type="button" class="btn btn-primary" data-toggle="modal" data-target="#Modalrup" onclick="choicerup()">Подобрать РУП</button>'

# print(find_value_with_substring(html, substring))

# wb = load_workbook(
#     filename='flask\static\IMI rasp ochno 2 polug 2022-2023_28.02 (1).xlsx')
# sheets_names = wb.sheetnames

# groups = {
#     "Б-ИВТ-21-1": "09.03.01|7618|ИМИ-Б-ИВТ-21-1|5954",
#     "Б-М-21": "01.03.01|7468|ИМИ-Б-М-21|5996",
#     "Б-ФИИТ-21": "02.03.02|7471|ИМИ-Б-ФИИТ-21|5998",
#     "Б-ИВТ-21-2": "09.03.01|7619|ИМИ-Б-ИВТ-21-2|5999",
#     "Б-ИТСС-21": "11.03.02|7467|ИМИ-Б-ИТСС-21|6003",
#     "Б-МПО-21": "44.03.01|7469|ИМИ-Б-МПО-21|6004",
#     "Б-ПОИМ-21": "44.03.05|7470|ИМИ-Б-ПОИМ-21|6005",
#     "Б-ПИГМУ-21": "09.03.03|7856|ИМИ-Б-ПИГМУ-21|6406",
#     "Б-ПИЭ-21": "09.03.03|7855|ИМИ-Б-ПИЭ-21|6407",
#     "Б-ПМИ-21": "01.03.02|7623|ИМИ-Б-ПМИ-21|6724",
# }

# group_data = groups.get("Б-ПИЭ-21")

# print(group_data)

# # цикл по листам excel-файла
# # for sh in sheets_names:
# #     wb.active = sheets_names.index(sh)
# #     ws = wb.active
# #     schedule = {}
# #     weekday = ""
# #     course = ws.cell(row=2, column=1).value
# #     year_and_semestr = ws.cell(row=1, column=1).value
# #     if course is None and year_and_semestr is None:
# #         continue
# #     print(course, year_and_semestr)

# #     for row in ws.iter_rows():
# #         if row[0].value == "Суббота":
# #             max_row = row[0].row

# #     # цикл по всем группам
# #     for i in range(3, ws.max_column, 4):
# #         group_name = ws.cell(row=4, column=i).value
# #         if group_name != "**" and group_name != "*":
# #             # цикл по занятиям 1-ой группы
# #             for j in range(6, max_row + 1):
# #                 lesson = {}
# #                 if ws.cell(row=j, column=i).value is not None:
# #                     if ws.cell(row=j, column=1).value is not None:
# #                         weekday = ws.cell(row=j, column=1).value

# #                     lesson = {
# #                         "номер пары": j - 5,
# #                         "день недели": (weekday, ""),
# #                         "временной отрезок": (ws.cell(row=j, column=2).value, ""),
# #                         "название дисциплины": (ws.cell(row=j, column=3).value, ""),
# #                         "ФИО преподавателя": (ws.cell(row=j, column=4).value, ""),
# #                         "вид деятельности": (ws.cell(row=j, column=5).value, ""),
# #                         "номер аудитории": (ws.cell(row=j, column=6).value, ""),
# #                     }
# #                     schedule.setdefault(
# #                         group_name, []).append(lesson)
